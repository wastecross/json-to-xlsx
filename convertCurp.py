# Code to create a xslx file from json (CURP ==> RENAPO).
# Install openpyxl pip3 install openpyxl


import json
import openpyxl
from openpyxl import Workbook


folder = 'folder_name'
file = 'renapo'
file_json = f'./json/{folder}/{file}.json'
file_xslx = f'./excel/{folder}/{file}.xlsx'


def set_id():
    if 'id' in curpRecord:
        ws_01.cell(row, 1, curpRecord["id"])
    else:
        ws_01.cell(row, 1, "")


def set_createdAt():
    if 'createdAt' in curpRecord:
        ws_01.cell(row, 2, curpRecord["createdAt"])
    else:
        ws_01.cell(row, 2, "")


def set_uuid():
    if 'uuid' in curpRecord:
        ws_01.cell(row, 3, curpRecord["uuid"])
    else:
        ws_01.cell(row, 3, "")


if __name__ == '__main__':

    json_data = {}

    with open(file_json) as json_file:
        json_data = json.load(json_file)

    wb = Workbook()

    # Grab the active worksheet
    ws_01 = wb.active

    # Set the title of the worksheet
    ws_01.title = 'RENAPO Records'

    # Set first row
    ws_01.cell(1, 1, "ID")
    ws_01.cell(1, 2, "Created At")
    ws_01.cell(1, 3, "UUID")

    row = 1
    for curpRecord in json_data.get("verifications"):
        row += 1
        set_id()
        set_createdAt()
        set_uuid()

    # Save it in an Excel file
    wb.save(file_xslx)
