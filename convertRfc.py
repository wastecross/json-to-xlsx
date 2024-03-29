# Code to create a xslx file from json (INE).
# Install openpyxl pip3 install openpyxl


import json
import openpyxl
from openpyxl import Workbook


folder = 'folder_name'
file = 'rfc'
fileJson = f'./json/{folder}/{file}.json'
fileXslx = f'./excel/{folder}/{file}.xlsx'


def set_id():
    if 'id' in rfcRecord:
        ws_01.cell(row, 1, rfcRecord["id"])
    else:
        ws_01.cell(row, 1, "")


def set_createdAt():
    if 'createdAt' in rfcRecord:
        ws_01.cell(row, 2, rfcRecord["createdAt"])
    else:
        ws_01.cell(row, 2, "")


if __name__ == '__main__':

    json_data = {}

    with open(fileJson) as json_file:
        json_data = json.load(json_file)

    wb = Workbook()

    # Grab the active worksheet
    ws_01 = wb.active

    # Set the title of the worksheet
    ws_01.title = 'RFC Records'

    # Set first row
    ws_01.cell(1, 1, "ID")
    ws_01.cell(1, 2, "Created At")

    row = 1
    for rfcRecord in json_data.get("verifications"):
        row += 1
        set_id()
        set_createdAt()

    # Save it in an Excel file
    wb.save(fileXslx)
