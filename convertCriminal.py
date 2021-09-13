# Code to create a xslx file from json (Criminal Record).
# Install openpyxl pip3 install openpyxl


import json
import openpyxl
from openpyxl import Workbook


folder = 'folder_name'
file = 'criminal'
fileJson = f'./json/{folder}/{file}.json'
fileXslx = f'./excel/{folder}/{file}.xlsx'


def set_id():
    if 'id' in criminalRecord:
        ws_01.cell(row, 1, criminalRecord["id"])
    else:
        ws_01.cell(row, 1, "")


def set_createdAt():
    if 'createdAt' in criminalRecord:
        ws_01.cell(row, 2, criminalRecord["createdAt"])
    else:
        ws_01.cell(row, 2, "")


def set_uuid():
    if 'uuid' in criminalRecord:
        ws_01.cell(row, 3, criminalRecord["uuid"])
    else:
        ws_01.cell(row, 3, "")


if __name__ == '__main__':

    json_data = {}

    with open(fileJson) as json_file:
        json_data = json.load(json_file)

    wb = Workbook()

    # Grab the active worksheet
    ws_01 = wb.active

    # Set the title of the worksheet
    ws_01.title = 'Criminal Records'

    # Set first row
    ws_01.cell(1, 1, "ID")
    ws_01.cell(1, 2, "Created At")
    ws_01.cell(1, 3, 'UUID')

    row = 1
    for criminalRecord in json_data.get("verifications"):
        row += 1
        set_id()
        set_createdAt()
        set_uuid()

    # Save it in an Excel file
    wb.save(fileXslx)
