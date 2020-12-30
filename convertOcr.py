# Code to create a xslx file from json (OCR).
# Install openpyxl pip3 install openpyxl


import json
import openpyxl
from openpyxl import Workbook


folder = 'Caliente26Nov-27Dic'
fileJson = f'./json/{folder}/ocr.json'
fileXslx = f'./excel/{folder}/ocr.xlsx'


def set_name():
    if 'name' in ocr:
        ws_01.cell(row, 3, ocr["name"])
    else:
        ws_01.cell(row, 3, "")


def set_invoiceDate():
    if 'invoiceDate' in ocr:
        ws_01.cell(row, 4, ocr["invoiceDate"])
    else:
        ws_01.cell(row, 4, "")


def set_address():
    if 'address' in ocr:
        ws_01.cell(row, 5, ocr["address"])
    else:
        ws_01.cell(row, 5, "")


if __name__ == '__main__':

    json_data = {}

    with open(fileJson) as json_file:
        json_data = json.load(json_file)

    wb = Workbook()

    # Grab the active worksheet
    ws_01 = wb.active

    # Set the title of the worksheet
    ws_01.title = 'Ocr Records'

    # Set first row
    ws_01.cell(1, 1, "Scan ID")
    ws_01.cell(1, 2, "Start Date UTC")
    ws_01.cell(1, 3, "Name")
    ws_01.cell(1, 4, "Invoice Date")
    ws_01.cell(1, 5, "Address")

    row = 1
    for ocr in json_data.get("scans"):
        row += 1
        ws_01.cell(row, 1, ocr["scanId"])
        ws_01.cell(row, 2, ocr["startDateUtc"])
        set_name()
        set_invoiceDate()
        set_address()

    # Save it in an Excel file
    wb.save(fileXslx)
