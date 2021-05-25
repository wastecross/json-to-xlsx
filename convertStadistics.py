# Code to create a xslx file from json (INE).
# Install openpyxl pip3 install openpyxl


import json
import openpyxl
from openpyxl import Workbook


folder = 'folder_name'
fileJson = f'./json/{folder}/stadistics.json'
fileXslx = f'./excel/{folder}/stadistics.xlsx'


def set_id():
    if 'id' in stadisticsRecord:
        ws_01.cell(row, 1, stadisticsRecord["id"])
    else:
        ws_01.cell(row, 1, "")


def set_createdAt():
    if 'createdAt' in stadisticsRecord:
        ws_01.cell(row, 2, stadisticsRecord["createdAt"])
    else:
        ws_01.cell(row, 2, "")


def set_data():
    if 'data' in stadisticsRecord:
        for attr, value in stadisticsRecord["data"].items():
            if attr == 'successChecks':
                ws_01.cell(row, 3, value)
            if attr == 'warningChecks':
                ws_01.cell(row, 4, value)
            if attr == 'failedChecks':
                ws_01.cell(row, 5, value)
            if attr == 'globalResult':
                ws_01.cell(row, 6, value)
    else:
        ws_01.cell(row, 3, "")
        ws_01.cell(row, 4, "")
        ws_01.cell(row, 5, "")
        ws_01.cell(row, 6, "")


if __name__ == '__main__':

    json_data = {}

    with open(fileJson) as json_file:
        json_data = json.load(json_file)

    wb = Workbook()

    # Grab the active worksheet
    ws_01 = wb.active

    # Set the title of the worksheet
    ws_01.title = 'INE Records'

    # Set first row
    ws_01.cell(1, 1, "ID")
    ws_01.cell(1, 2, "Created At")
    ws_01.cell(1, 3, "successChecks")
    ws_01.cell(1, 4, "warningChecks")
    ws_01.cell(1, 5, "failedChecks")
    ws_01.cell(1, 6, "globalResult")

    row = 1
    for stadisticsRecord in json_data.get("verifications"):
        row += 1
        set_id()
        set_createdAt()
        set_data()

    # Save it in an Excel file
    wb.save(fileXslx)
