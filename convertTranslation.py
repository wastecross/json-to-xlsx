# Code to create a xslx file from json (INE).
# Install openpyxl pip3 install openpyxl


import json
import openpyxl
from openpyxl import Workbook


folder = 'translations'
file = 'api'
fileJson = f'./json/{folder}/{file}.json'
fileXslx = f'./excel/{folder}/{file}.xlsx'


def set_language_english():
    if 'en' in translationRecord:
        ws_01.cell(row, 1, translationRecord["en"])
    else:
        ws_01.cell(row, 1, "")


def set_language_spanish():
    if 'es' in translationRecord:
        ws_01.cell(row, 2, translationRecord["es"])
    else:
        ws_01.cell(row, 2, "")


if __name__ == '__main__':

    json_data = {}

    with open(fileJson) as json_file:
        json_data = json.load(json_file)

    wb = Workbook()

    # Grab the active worksheet
    ws_01 = wb.active

    # Set the title of the worksheet
    ws_01.title = 'Translation'

    # Set first row
    ws_01.cell(1, 1, "English")
    ws_01.cell(1, 2, "Spanish")

    row = 1
    for translationRecord in json_data.get("languages"):
        row += 1
        set_language_english()

    row = 1
    for translationRecord in json_data.get("languages"):
        row += 1
        set_language_spanish()

    # Save it in an Excel file
    wb.save(fileXslx)
