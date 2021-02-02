# Code to create a xslx file from json (ID).
# Install openpyxl pip3 install openpyxl


import json
import openpyxl
from openpyxl import Workbook


folder = 'folder_name'
fileJson = f'./json/{folder}/id.json'
fileXslx = f'./excel/{folder}/id.xlsx'


def set_id():
    if 'id' in idRecord:
        ws_01.cell(row, 1, idRecord["id"])
    else:
        ws_01.cell(row, 1, "")


def set_checkTime():
    if 'checkTime' in idRecord:
        ws_01.cell(row, 4, idRecord["checkTime"])
    else:
        ws_01.cell(row, 4, "")


def set_documentType():
    if 'documentType' in idRecord:
        ws_01.cell(row, 5, idRecord["documentType"])
    else:
        ws_01.cell(row, 5, "")


def set_countryCode():
    if 'countryCode' in idRecord:
        ws_01.cell(row, 6, idRecord["countryCode"])
    else:
        ws_01.cell(row, 6, "")


def set_spentCredits():
    if 'spentCredits' in idRecord:
        ws_01.cell(row, 7, idRecord["spentCredits"])
    else:
        ws_01.cell(row, 7, "")


def set_totalChecks():
    if 'totalChecks' in idRecord:
        ws_01.cell(row, 9, idRecord["totalChecks"])
    else:
        ws_01.cell(row, 9, "")


def set_successChecks():
    if 'successChecks' in idRecord:
        ws_01.cell(row, 10, idRecord["successChecks"])
    else:
        ws_01.cell(row, 10, "")


def set_warningChecks():
    if 'warningChecks' in idRecord:
        ws_01.cell(row, 11, idRecord["warningChecks"])
    else:
        ws_01.cell(row, 11, "")


def set_failedChecks():
    if 'failedChecks' in idRecord:
        ws_01.cell(row, 12, idRecord["failedChecks"])
    else:
        ws_01.cell(row, 12, "")


def set_faceVerification():
    if 'faceVerification' in idRecord:
        ws_01.cell(row, 13, idRecord["faceVerification"])
    else:
        ws_01.cell(row, 13, "")


if __name__ == '__main__':

    json_data = {}

    with open(fileJson) as json_file:
        json_data = json.load(json_file)

    wb = Workbook()

    # Grab the active worksheet
    ws_01 = wb.active

    # Set the title of the worksheet
    ws_01.title = 'ID Records'

    # Set first row
    ws_01.cell(1, 1, "ID")
    ws_01.cell(1, 2, "Identifier")
    ws_01.cell(1, 3, "Start Date Utc")
    ws_01.cell(1, 4, "Check Time")
    ws_01.cell(1, 5, "Document Type")
    ws_01.cell(1, 6, "Country Code")
    ws_01.cell(1, 7, "Spent Credits")
    ws_01.cell(1, 8, "Status")
    ws_01.cell(1, 9, "Total Checks")
    ws_01.cell(1, 10, "Success Checks")
    ws_01.cell(1, 11, "Warning Checks")
    ws_01.cell(1, 12, "Failed Checks")
    ws_01.cell(1, 13, "Face Verification")

    row = 1
    for idRecord in json_data.get("verifications"):
        row += 1
        set_id()
        ws_01.cell(row, 2, idRecord["identifier"])
        ws_01.cell(row, 3, idRecord["startDateUtc"])
        set_checkTime()
        set_documentType()
        set_countryCode()
        set_spentCredits()
        ws_01.cell(row, 8, idRecord["status"])
        set_totalChecks()
        set_successChecks()
        set_warningChecks()
        set_failedChecks()
        set_faceVerification()

    # Save it in an Excel file
    wb.save(fileXslx)
